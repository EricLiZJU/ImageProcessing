import cv2
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook
import sys

original_stdout = sys.stdout

clusteredmaincolorfilepath = 'clusteredmaincolor.xlsx'
clusteredauxiliarycolorfilepath = 'clusteredauxiliarycolor.xlsx'

workbook1 = load_workbook(filename=clusteredmaincolorfilepath)
workbook2 = load_workbook(filename=clusteredauxiliarycolorfilepath)
workbook3 = Workbook()
workbook4 = Workbook()


sheet1 = workbook1.active
sheet2 = workbook2.active
sheet3 = workbook3.active
sheet4 = workbook4.active

maincolors = np.empty((1, 1, 3), int)
auxiliarycolors = np.empty((1, 1, 3), int)

for row in sheet1.iter_rows(values_only=True):
    row = list(row)
    row = row[0].split()
    row = [float(i) for i in row]
    row = [round(i) for i in row]
    BGRvalues = np.array([[row]], dtype=np.uint8)
    RGBvalues = cv2.cvtColor(BGRvalues, cv2.COLOR_BGR2RGB)
    HSVvalues = cv2.cvtColor(RGBvalues, cv2.COLOR_RGB2HSV)
    maincolors = np.append(maincolors, HSVvalues, axis=0)

maincolors = np.delete(maincolors, 0, axis=0)
print(maincolors)

for row in sheet2.iter_rows(values_only=True):
    row = list(row)
    row = row[0].split()
    row = [float(i) for i in row]
    row = [round(i) for i in row]
    BGRvalues = np.array([[row]], dtype=np.uint8)
    RGBvalues = cv2.cvtColor(BGRvalues, cv2.COLOR_BGR2RGB)
    HSVvalues = cv2.cvtColor(RGBvalues, cv2.COLOR_RGB2HSV)
    auxiliarycolors = np.append(auxiliarycolors, HSVvalues, axis=0)

auxiliarycolors = np.delete(auxiliarycolors, 0, axis=0)
print(auxiliarycolors)

with open('HSVvalue.txt', 'w') as file:
    sys.stdout = file
    print("maincolors:\n")
    print(maincolors)
    print("auxiliarycolors:\n")
    print(auxiliarycolors)
sys.stdout = original_stdout

# 提取色相 (H)、饱和度 (S) 和明度 (V)
H1 = maincolors[:, :, 0]  # 色相
S1 = maincolors[:, :, 1]  # 饱和度
V1 = maincolors[:, :, 2]  # 明度

H2 = auxiliarycolors[:, :, 0]  # 色相
S2 = auxiliarycolors[:, :, 1]  # 饱和度
V2 = auxiliarycolors[:, :, 2]  # 明度

# 将色相转换为弧度，因为极坐标图中的角度是弧度
theta1 = np.deg2rad(H1)
theta2 = np.deg2rad(H2)

# 创建极坐标图
fig = plt.figure()
ax1 = fig.add_subplot(111, projection='polar')

# HSV颜色的转换：matplotlib支持RGB，所以我们需要从HSV转换到RGB
for i in range(len(maincolors)):
    color = plt.cm.hsv(H1[i] / 360)  # HSV to RGB
    ax1.scatter(theta1[i], S1[i], s=100,
                color='orange',
                #label=f'H={H1[i]}, S={S1[i]}, V={V1[i]}',
                marker='.')

for i in range(len(auxiliarycolors)):
    color = plt.cm.hsv(H2[i] / 360)  # HSV to RGB
    ax1.scatter(theta2[i], S2[i], s=100,
                color='blue',
                #label=f'H={H2[i]}, S={S2[i]}, V={V2[i]}',
                marker='.')

for i in range(len(maincolors)):
    color = plt.cm.hsv(H1[i] / 360)  # HSV to RGB
    ax1.scatter(theta1[i], V1[i], s=100,
                color='orange',
                #label=f'H={H1[i]}, S={S1[i]}, V={V1[i]}',
                marker='1')

for i in range(len(auxiliarycolors)):
    color = plt.cm.hsv(H2[i] / 360)  # HSV to RGB
    ax1.scatter(theta2[i], V2[i], s=100,
                color='blue',
                #label=f'H={H2[i]}, S={S2[i]}, V={V2[i]}',
                marker='1')

# 设置图形标签
ax1.set_title('HSV Color Polar Plot', va='bottom')
ax1.set_rticks([0.2, 0.4, 0.6, 0.8, 1.0])  # 饱和度刻度
ax1.set_rlabel_position(-22.5)  # 标签显示在合适位置

# 显示图例
#plt.legend(loc='upper right', bbox_to_anchor=(1.2, 1.1))

# 显示图形
plt.show()


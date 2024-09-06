import cv2
import numpy as np
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
import pandas as pd
import itertools
from itertools import combinations
from openpyxl import load_workbook
from openpyxl import Workbook
import sys

original_stdout = sys.stdout

def int_list_to_string(int_list):
    return ' '.join(str(num) for num in int_list)


# 计算候选项集的支持度
def calculate_support(transactions, itemsets):
    support = {}
    for itemset in itemsets:
        support_count = sum(1 for transaction in transactions if set(itemset).issubset(transaction))
        support[itemset] = support_count
    return support


# 生成候选项集
def generate_candidates(prev_freq_itemsets, k):
    candidates = []
    prev_itemsets_list = list(prev_freq_itemsets.keys())
    for i in range(len(prev_itemsets_list)):
        for j in range(i + 1, len(prev_itemsets_list)):
            # 合并两个频繁项集，生成候选项集
            candidate = tuple(sorted(set(prev_itemsets_list[i]) | set(prev_itemsets_list[j])))
            if len(candidate) == k:
                candidates.append(candidate)
    return set(candidates)

# 计算支持度
def calculate_support_value(itemset, transactions):
    return sum(1 for transaction in transactions if set(itemset).issubset(transaction)) / len(transactions)

# 筛选支持度符合要求的频繁项集
def filter_frequent_itemsets(support, min_support):
    return {itemset: count for itemset, count in support.items() if count >= min_support}

# 生成关联规则并计算置信度
def generate_association_rules(freq_itemsets, transactions, min_confidence):
    rules = []
    for itemset in freq_itemsets:
        if len(itemset) > 1:
            # 对于每个频繁项集，生成所有可能的前件 A 和后件 B
            for i in range(1, len(itemset)):
                for antecedent in combinations(itemset, i):
                    consequent = tuple(set(itemset) - set(antecedent))
                    antecedent_support = calculate_support_value(antecedent, transactions)
                    rule_support = calculate_support_value(itemset, transactions)
                    confidence = rule_support / antecedent_support
                    if confidence >= min_confidence:
                        rules.append((antecedent, consequent, confidence))
    return rules

# Apriori算法实现
def apriori(transactions, min_support, min_confidence):
    # 生成初始候选项集
    transactions = [tuple(transaction) for transaction in transactions]
    items = set(item for transaction in transactions for item in transaction)
    itemsets = [item for item in items]
    itemsets = tuple(itemsets)

    print(itemsets)

    # 计算频繁项集
    freq_itemsets = {}
    k = 1
    while itemsets:
        support = calculate_support(transactions, itemsets)
        frequent_itemsets = filter_frequent_itemsets(support, min_support)

        if not frequent_itemsets:
            break

        freq_itemsets.update(frequent_itemsets)
        k += 1
        # 生成下一个级别的候选项集
        itemsets = generate_candidates(frequent_itemsets, k)

    # 生成关联规则
    rules = generate_association_rules(freq_itemsets.keys(), transactions, min_confidence)

    return freq_itemsets, rules

colordatapath = ('colordata.xlsx')
clusteredmaincolorpath = ('clusteredmaincolor.xlsx')
clusteredauxiliarypath = ('clusteredauxiliarycolor.xlsx')

workbook1 = load_workbook(filename=colordatapath)
workbook2 = load_workbook(filename=clusteredmaincolorpath)
workbook3 = load_workbook(filename=clusteredauxiliarypath)

sheet1 = workbook1.active
sheet2 = workbook2.active
sheet3 = workbook3.active
colordatas = []
maincolordatas = []
auxiliarycolordatas = []
datasforapriori = []
maincolorclasstoadd = []
auxiliarycolorclasstoadd = []
maincolorclass = []
auxiliarycolorclass = []
finaldata = []

delta = list(range(-10, 11))

for row in sheet1.iter_rows(values_only=True):
    row = list(row)
    row = [row[i] for i in [1, 3, 5, 7, 9, 11]]
    colordatas.append(row)

colordatas.pop(0)

for row in sheet2.iter_rows(values_only=True):
    maincolorclasstoadd = []
    row = list(row)
    row = row[0].split()
    row = [round(float(i)) for i in row]
    combinations1 = list(itertools.product(delta, repeat=len(row)))
    result = []
    for comb in combinations1:
        new_array = [row[i] + comb[i] for i in range(len(row))]
        new_array = int_list_to_string(new_array)
        maincolorclasstoadd.append(new_array)
        maincolordatas.append(new_array)
    maincolorclass.append([maincolorclasstoadd])

for row in sheet3.iter_rows(values_only=True):
    auxiliarycolorclasstoadd = []
    row = list(row)
    row = row[0].split()
    row = [round(float(i)) for i in row]
    combinations1 = list(itertools.product(delta, repeat=len(row)))
    result = []
    for comb in combinations1:
        new_array = [row[i] + comb[i] for i in range(len(row))]
        new_array = int_list_to_string(new_array)
        auxiliarycolorclasstoadd.append(new_array)
        auxiliarycolordatas.append(new_array)
    auxiliarycolorclass.append([auxiliarycolorclasstoadd])

for i in colordatas:
    toadd = []
    for j in i:
        if (j in maincolordatas) or (j in auxiliarycolordatas):
            toadd.append(j)

    if len(toadd) > 1:
        datasforapriori.append([toadd])
        toadd = []

#print(maincolorclass[1])
for i in datasforapriori:
    for k in i:
        for j in k:
            if j in maincolorclass[0][0]:
                toadd.append('A')
                continue
            if j in maincolorclass[1][0]:
                toadd.append('B')
                continue
            if j in maincolorclass[2][0]:
                toadd.append('C')
                continue
            if j in maincolorclass[3][0]:
                toadd.append('D')
                continue
            if j in maincolorclass[4][0]:
                toadd.append('E')
                continue
            if j in auxiliarycolorclass[0][0]:
                toadd.append('a')
                continue
            if j in auxiliarycolorclass[1][0]:
                toadd.append('b')
                continue
            if j in auxiliarycolorclass[2][0]:
                toadd.append('c')
                continue
            if j in auxiliarycolorclass[3][0]:
                toadd.append('d')
                continue
            if j in auxiliarycolorclass[4][0]:
                toadd.append('e')
                continue
            if j in auxiliarycolorclass[5][0]:
                toadd.append('f')
                continue
            if j in auxiliarycolorclass[6][0]:
                toadd.append('g')
                continue
            if j in auxiliarycolorclass[7][0]:
                toadd.append('h')
                continue
            if j in auxiliarycolorclass[8][0]:
                toadd.append('i')
                continue
            if j in auxiliarycolorclass[9][0]:
                toadd.append('j')
                continue
    #print(toadd)
    finaldata.append(toadd)
    toadd = []


min_support = 2
min_confidence = 0.6

freq_itemsets, rules = apriori(finaldata, min_support, min_confidence)

with open('output.txt', 'w') as file:
    sys.stdout = file

    # 打印频繁项集和支持度
    print("频繁项集和支持度:")
    for itemset, support_count in freq_itemsets.items():
        support_value = support_count / len(finaldata)
        print(f"{set(itemset)}: 支持度 = {support_value}")

    # 打印关联规则和置信度
    print("\n关联规则和置信度:")
    for antecedent, consequent, confidence in rules:
        print(f"{set(antecedent)} -> {set(consequent)}: 置信度 = {confidence}")

sys.stdout = original_stdout







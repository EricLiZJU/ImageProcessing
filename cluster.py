import cv2
import numpy as np
import os
from openpyxl import Workbook
import matplotlib.pyplot as plt

wb1 = Workbook()
ws1 = wb1.active
wb2 = Workbook()
ws2 = wb2.active
wb3 = Workbook()
ws3 = wb3.active

def int_list_to_string(int_list):
    return ' '.join(str(num) for num in int_list)

def max_index(arr):
    return arr.index(max(arr))


tabledata = [
    ['filename', 'color1', 'color1_ratio',
     'color2', 'color2_ratio',
     'color3', 'color3_ratio',
     'color4', 'color4_ratio',
     'color5', 'color5_ratio',
     'color6', 'color6_ratio',
     'color7', 'color7_ratio'],
]
maincolorlist = [
    ['maincolor'],
]
auxiliarycolorlist = [
    ['auxiliarycolor'],
]

folder_path = 'FiltedImage'
file_names = os.listdir(folder_path)
file_names = [file_name for file_name in file_names if file_name != '.DS_Store']
count = 1

for file_name in file_names:
    print(count)
    image = cv2.imread('FiltedImage/' + file_name)
    cv2.imshow("input", image)
    h, w, ch = image.shape

    # 构建图像数据
    data = image.reshape((-1, 3))
    data = np.float32(data)

    # 图像分割
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 10, 1.0)
    num_clusters = 7
    ret, label, center = cv2.kmeans(data, num_clusters, None, criteria, num_clusters, cv2.KMEANS_RANDOM_CENTERS)

    # 生成主色彩条形卡片
    card = np.zeros((50, w, 3), dtype=np.uint8)
    clusters = np.zeros([7], dtype=np.int32)
    for i in range(len(label)):
        clusters[label[i][0]] += 1
    # 计算各类别像素的比率
    clusters = np.float32(clusters) / float(h*w)
    center = np.int32(center)
    x_offset = 0
    ratio = []
    total = 0
    strcenter = []
    for c in range(num_clusters):
        dx = int(clusters[c] * w)
        ratio.append(dx)
        b = center[c][0]
        g = center[c][1]
        r = center[c][2]
        cv2.rectangle(card, (x_offset, 0), (x_offset+dx, 50),
                     (int(b),int(g),int(r)), -1)
        x_offset += dx
    total = sum(ratio)
    maincolorindex = max_index(ratio)


    # 聚类后图片显示
    centers = np.uint8(center)
    res = centers[label.flatten()]
    ds = res.reshape((image.shape))
    #ds = cv.cvtColor(ds,)

    for i in range(7):
        ratio[i] = ratio[i] / total
        ratio[i] = str(ratio[i])
    for i in range(7):
        intcenter = center[i].tolist()
        strcenter.append(int_list_to_string(intcenter))
    print(strcenter)

    for i in range(7):
        if i == maincolorindex:
            maincolortoadd = [strcenter[i]]
            maincolorlist.append(maincolortoadd)
        else:
            auxiliarycolortoadd = [strcenter[i]]
            auxiliarycolorlist.append(auxiliarycolortoadd)

    newdata = [file_name, strcenter[0], ratio[0],
               strcenter[1], ratio[1],
               strcenter[2], ratio[2],
               strcenter[3], ratio[3],
               strcenter[4], ratio[4],
               strcenter[5], ratio[5],
               strcenter[6], ratio[6]]
    tabledata.append(newdata)
    cv2.imwrite('ColorTable/' + file_name, card)
    #cv2.imwrite('TestClusteredImage/' + file_name, ds)
    #cv2.imwrite('TestClusteredImage/card.jpeg', card)
    count = count + 1

for row in tabledata:
    ws1.append(row)
for row in maincolorlist:
    ws2.append(row)
for row in auxiliarycolorlist:
    ws3.append(row)

wb1.save('colordata.xlsx')
wb2.save('maincolorlist.xlsx')
wb3.save('auxiliarycolorlist.xlsx')
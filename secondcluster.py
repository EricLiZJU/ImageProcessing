import cv2
import numpy as np
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

def float_list_to_string(float_list):
    return ' '.join(str(num) for num in float_list)

maincolorfilepath = 'maincolorlist.xlsx'
auxiliarycolorfilepath = 'auxiliarycolorlist.xlsx'

workbook1 = load_workbook(filename=maincolorfilepath)
workbook2 = load_workbook(filename=auxiliarycolorfilepath)
workbook3 = Workbook()
workbook4 = Workbook()

sheet1 = workbook1.active
sheet2 = workbook2.active
sheet3 = workbook3.active
sheet4 = workbook4.active

maincolorstoadd = np.array([])
auxiliarycolorstoadd = np.array([])
maincolors = np.empty((1, 3), int)
auxiliarycolors = np.array([[1, 1, 1]])

clusteredmaincolorlist = [
    ['clusteredmaincolor'],
]
clusteredauxiliarycolorlist = [
    ['clusteredauxiliarycolor'],
]

for row in sheet1.iter_rows(values_only=True):
    row = list(row)
    row = row[0].split()

    if row[0].isdigit():
        RGBvalues = np.array(row)
        RGBvalues = RGBvalues.astype(int)
        RGBvalues = np.array([RGBvalues])
        #print(np.shape(RGBvalues))
        maincolors = np.append(maincolors, RGBvalues, axis=0)
        #print(maincolors)

#print(maincolors)

for row in sheet2.iter_rows(values_only=True):
    row = list(row)
    row = row[0].split()

    if row[0].isdigit():
        RGBvalues = np.array(row)
        RGBvalues = RGBvalues.astype(int)
        RGBvalues = np.array([RGBvalues])
        # print(np.shape(RGBvalues))
        auxiliarycolors = np.append(auxiliarycolors, RGBvalues, axis=0)
        # print(maincolors)

#print(auxiliarycolors)


# 聚类生成主色
# 创建 KMeans 模型，设置聚类数量为 5
kmeans1 = KMeans(n_clusters=5, random_state=0)

# 训练模型
kmeans1.fit(maincolors)

# 获取聚类的中心点
main_centroids = kmeans1.cluster_centers_
main_centroids = main_centroids.tolist()
# 输出聚类中心点
print("main_centroids:\n", main_centroids)

# 获取每个样本的聚类标签
main_labels = kmeans1.labels_

for c in range(5):
    image1 = np.zeros((400, 400, 3), dtype=np.uint8)
    start_point = (0, 0)
    end_point = (400, 400)
    color = (main_centroids[c][0], main_centroids[c][1], main_centroids[c][2])
    cv2.rectangle(image1, start_point, end_point, color, -1)
    cv2.imwrite('MainColor/' + str(c) + '.jpeg', image1)

# 聚类生成辅色
# 创建 KMeans 模型，设置聚类数量为 10
kmeans2 = KMeans(n_clusters=10, random_state=0)

# 训练模型
kmeans2.fit(auxiliarycolors)

# 获取聚类的中心点
auxiliary_centroids = kmeans2.cluster_centers_
auxiliary_centroids = auxiliary_centroids.tolist()

# 输出聚类中心点
print("auxiliary_centroids:\n", auxiliary_centroids)

# 获取每个样本的聚类标签
auxiliary_labels = kmeans2.labels_

for c in range(10):
    image2 = np.zeros((400, 400, 3), dtype=np.uint8)
    start_point = (0, 0)
    end_point = (400, 400)
    color = (auxiliary_centroids[c][0], auxiliary_centroids[c][1], auxiliary_centroids[c][2])
    cv2.rectangle(image2, start_point, end_point, color, -1)
    cv2.imwrite('AuxiliaryColor/' + str(c) + '.jpeg', image2)

for row in main_centroids:
    row = float_list_to_string(row)
    sheet3.append([row])

for row in auxiliary_centroids:
    row = float_list_to_string(row)
    sheet4.append([row])

workbook3.save('clusteredmaincolor.xlsx')
workbook4.save('clusteredauxiliarycolor.xlsx')